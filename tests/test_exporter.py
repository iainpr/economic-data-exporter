from datetime import UTC, datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from economic_data_exporter.models import FailureRecord, SeriesMetadata, SourceResult
from economic_data_exporter.services.exporter import export_workbook, safe_sheet_name
from economic_data_exporter.services.ingestion import canonicalize_observations


def test_excel_creation_readme_data_failures(tmp_path: Path, sample_result) -> None:
    output = tmp_path / "result.xlsx"
    failure = FailureRecord("FRED", "BAD", "", "ParsingError", "Malformed")
    path = export_workbook(output, [sample_result], [failure])
    workbook = load_workbook(path, read_only=True, data_only=False)
    assert {"README", "Data", "Failures"}.issubset(workbook.sheetnames)
    assert workbook["Data"].max_row == 4
    assert workbook["README"].max_row > 5
    workbook.close()


def test_formula_injection_guarded_on_stringdtype_columns(tmp_path: Path, sample_request) -> None:
    """series_name etc. arrive as pandas StringDtype after ingestion, not object dtype.

    _safe_data_for_excel must select by content, not dtype == "object", or the
    apostrophe guard silently does nothing for these columns.
    """

    retrieved = datetime(2026, 8, 5, 12, 0, tzinfo=UTC)
    raw = pd.DataFrame(
        {
            "date": ["2020-01-01"],
            "value": [1.0],
            "series_id": ["FXUSDCAD"],
            "series_name": ['=HYPERLINK("http://evil","click")'],
            "geography": ["Canada"],
            "frequency": ["Daily"],
            "units": ["CAD"],
            "retrieved_at": [retrieved.isoformat()],
            "source_url": ["https://example.com"],
        }
    )
    data = canonicalize_observations(raw, source="Bank of Canada")
    assert pd.api.types.is_string_dtype(data["series_name"])

    metadata = SeriesMetadata(
        source="Bank of Canada",
        series_id="FXUSDCAD",
        name="USD/CAD",
        geography="Canada",
        frequency="Daily",
        units="CAD",
        source_url="https://example.com",
    )
    result = SourceResult(
        request=sample_request, data=data, metadata=metadata, retrieved_at=retrieved
    )

    output = tmp_path / "injection.xlsx"
    path = export_workbook(output, [result], [])
    workbook = load_workbook(path, read_only=True, data_only=False)
    worksheet = workbook["Data"]
    header = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    column_index = header.index("series_name") + 1
    cell_value = worksheet.cell(row=2, column=column_index).value
    workbook.close()

    assert cell_value.startswith("'=")


def test_safe_sheet_names() -> None:
    used: set[str] = set()
    first = safe_sheet_name("a/b:c*?[]" * 5, used)
    second = safe_sheet_name("a/b:c*?[]" * 5, used)
    assert len(first) <= 31
    assert first != second
