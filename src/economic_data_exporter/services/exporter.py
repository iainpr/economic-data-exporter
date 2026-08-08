"""Deterministic, atomic Excel workbook export."""

from __future__ import annotations

import os
import re
import tempfile
from dataclasses import asdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Callable

import pandas as pd
from openpyxl import load_workbook

from economic_data_exporter import __version__
from economic_data_exporter.exceptions import ExportError
from economic_data_exporter.models import ExportFormatOptions, FailureRecord, SourceResult
from economic_data_exporter.services.preparation import prepare_export_data
from economic_data_exporter.services.provenance import run_fingerprint
from economic_data_exporter.utils.validation import validate_output_path

_INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")
_DANGEROUS_PREFIXES = ("=", "+", "-", "@")


def safe_excel_text(value: object) -> object:
    if not isinstance(value, str):
        return value
    return "'" + value if value.startswith(_DANGEROUS_PREFIXES) else value


def safe_sheet_name(name: str, existing: set[str]) -> str:
    cleaned = _INVALID_SHEET_CHARS.sub("_", name).strip().strip("'") or "Series"
    cleaned = cleaned[:31]
    candidate = cleaned
    suffix = 1
    while candidate.casefold() in {item.casefold() for item in existing}:
        suffix_text = f"_{suffix}"
        candidate = f"{cleaned[: 31 - len(suffix_text)]}{suffix_text}"
        suffix += 1
    existing.add(candidate)
    return candidate


def _format_options_rows(options: ExportFormatOptions, transformations: list[str]) -> list[dict[str, object]]:
    return [
        {"section": "Output", "field": "output_shape", "value": options.output_shape},
        {"section": "Output", "field": "selected_columns", "value": ", ".join(options.columns)},
        {"section": "Output", "field": "drop_missing_values", "value": options.drop_missing_values},
        {"section": "Output", "field": "remove_duplicate_rows", "value": options.remove_duplicate_rows},
        {"section": "Output", "field": "sort_by_date", "value": options.sort_by_date},
        {"section": "Output", "field": "freeze_header", "value": options.freeze_header},
        {"section": "Output", "field": "autofilter", "value": options.autofilter},
        {"section": "Output", "field": "auto_width", "value": options.auto_width},
        {"section": "Output", "field": "include_series_sheets", "value": options.include_series_sheets},
        {"section": "Output", "field": "transformations", "value": " | ".join(transformations)},
    ]


def _readme_rows(
    results: list[SourceResult],
    failures: list[FailureRecord],
    *,
    options: ExportFormatOptions,
    transformations: list[str],
    exported_rows: int,
    run_id: str,
) -> pd.DataFrame:
    rows: list[dict[str, object]] = [
        {"section": "Export", "field": "application_version", "value": __version__},
        {
            "section": "Export",
            "field": "export_timestamp_utc",
            "value": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
        },
        {"section": "Export", "field": "successful_series", "value": len(results)},
        {"section": "Export", "field": "failed_series", "value": len(failures)},
        {"section": "Export", "field": "exported_rows", "value": exported_rows},
        {"section": "Export", "field": "run_id", "value": run_id},
        {"section": "Export", "field": "idempotency", "value": "Repeated runs atomically replace the target workbook; no append/duplicate behavior is used."},
        *(_format_options_rows(options, transformations)),
        {
            "section": "Integrity",
            "field": "raw_data_policy",
            "value": "Provider observations are canonical; preview/clean/shape operations create a derived export view and do not mutate raw observations.",
        },
        {
            "section": "Integrity",
            "field": "excel_formula_safety",
            "value": "Text beginning with =, +, -, or @ is prefixed with an apostrophe to prevent formula execution.",
        },
        {
            "section": "FRED",
            "field": "required_notice",
            "value": "This product uses the FRED API but is not endorsed or certified by the Federal Reserve Bank of St. Louis.",
        },
    ]
    for index, result in enumerate(results, start=1):
        prefix = f"Series {index}"
        request = result.request
        metadata = result.metadata
        entries = {
            "source": metadata.source,
            "series_id": metadata.series_id,
            "series_name": metadata.name,
            "geography": metadata.geography or request.geography,
            "requested_start_date": request.start_date.isoformat(),
            "requested_end_date": request.end_date.isoformat(),
            "retrieved_at_utc": result.retrieved_at.replace(microsecond=0).isoformat(),
            "cache_status": "cache" if result.from_cache else "live",
            "frequency": metadata.frequency,
            "units": metadata.units,
            "source_url": metadata.source_url,
            "notes": metadata.notes,
            "warnings": " | ".join(result.warnings),
            "attribution": metadata.attribution,
            "license": metadata.license_text,
            "transformations": " | ".join(result.transformations) or "None",
        }
        rows.extend({"section": prefix, "field": key, "value": value} for key, value in entries.items())
    return pd.DataFrame(rows, columns=["section", "field", "value"])


def _safe_data_for_excel(data: pd.DataFrame) -> pd.DataFrame:
    frame = data.copy()
    for column in frame.columns:
        if pd.api.types.is_object_dtype(frame[column]) or pd.api.types.is_string_dtype(frame[column]):
            frame[column] = frame[column].map(safe_excel_text)
    return frame


def _style_sheet(ws, *, freeze_header: bool, autofilter: bool, auto_width: bool) -> None:
    if freeze_header and ws.max_row >= 1:
        ws.freeze_panes = "A2"
    if autofilter and ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions
    if auto_width:
        for column_cells in ws.iter_cols(min_row=1, max_row=min(ws.max_row, 1001)):
            max_length = 0
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            column_letter = column_cells[0].column_letter
            ws.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 42)


def export_workbook(
    output_path: Path,
    results: list[SourceResult],
    failures: list[FailureRecord],
    *,
    format_options: ExportFormatOptions | None = None,
    include_series_sheets: bool | None = None,
    progress: Callable[[str], None] | None = None,
) -> Path:
    final_path = validate_output_path(output_path)
    if not results:
        raise ExportError("No successful series are available to export.")
    options = format_options or ExportFormatOptions()
    if include_series_sheets is not None and include_series_sheets != options.include_series_sheets:
        options = ExportFormatOptions(
            output_shape=options.output_shape,
            columns=options.columns,
            drop_missing_values=options.drop_missing_values,
            remove_duplicate_rows=options.remove_duplicate_rows,
            sort_by_date=options.sort_by_date,
            freeze_header=options.freeze_header,
            autofilter=options.autofilter,
            auto_width=options.auto_width,
            include_series_sheets=include_series_sheets,
            preview_rows=options.preview_rows,
        )

    if progress:
        progress("Preparing cleaned export data")
    data, transformations = prepare_export_data(results, options)
    if len(data) + 1 > 1_048_576:
        raise ExportError("Data exceeds Excel's 1,048,576-row worksheet limit; narrow the selection.")
    if len(data.columns) > 16_384:
        raise ExportError("Data exceeds Excel's 16,384-column worksheet limit; narrow the selection or use long format.")

    run_id = run_fingerprint([result.request for result in results], options)
    readme = _readme_rows(
        results,
        failures,
        options=options,
        transformations=transformations,
        exported_rows=len(data),
        run_id=run_id,
    ).map(safe_excel_text)
    failures_frame = pd.DataFrame([asdict(record) for record in failures]) if failures else pd.DataFrame()
    if not failures_frame.empty:
        failures_frame = failures_frame.map(safe_excel_text)

    temp_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            prefix=f".{final_path.stem}.", suffix=".tmp.xlsx", dir=final_path.parent, delete=False
        ) as handle:
            temp_path = Path(handle.name)
        if progress:
            progress("Writing Excel workbook")
        with pd.ExcelWriter(temp_path, engine="openpyxl", datetime_format="yyyy-mm-dd") as writer:
            readme.to_excel(writer, sheet_name="README", index=False)
            _safe_data_for_excel(data).to_excel(writer, sheet_name="Data", index=False)
            if not failures_frame.empty:
                failures_frame.to_excel(writer, sheet_name="Failures", index=False)
            if options.include_series_sheets:
                used = {"README", "Data", "Failures"}
                for result in results:
                    subset = _safe_data_for_excel(result.normalized())
                    subset["date"] = pd.to_datetime(subset["date"], errors="coerce")
                    name = safe_sheet_name(f"{result.metadata.source}_{result.metadata.series_id}", used)
                    subset.to_excel(writer, sheet_name=name, index=False)

            # Style the in-memory workbook before the writer closes. This avoids
            # the old write -> reopen -> restyle -> save cycle.
            for worksheet in writer.book.worksheets:
                _style_sheet(
                    worksheet,
                    freeze_header=options.freeze_header,
                    autofilter=options.autofilter,
                    auto_width=options.auto_width,
                )

        if progress:
            progress("Validating workbook")
        workbook = load_workbook(temp_path, read_only=True, data_only=False)
        try:
            required = {"README", "Data"}
            if not required.issubset(workbook.sheetnames):
                raise ExportError("Workbook validation failed: required worksheets are missing.")
            if workbook["Data"].max_column != len(data.columns):
                raise ExportError("Workbook validation failed: Data worksheet schema is invalid.")
        finally:
            workbook.close()
        os.replace(temp_path, final_path)
        temp_path = None
        return final_path
    except ExportError:
        raise
    except Exception as exc:
        raise ExportError(f"Excel export failed: {exc}") from exc
    finally:
        if temp_path is not None:
            temp_path.unlink(missing_ok=True)
